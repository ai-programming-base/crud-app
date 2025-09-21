# blueprints/index_bp.py
from flask import Blueprint, render_template, request, send_file
from services import get_db, INDEX_FIELDS, login_required, logger, _cleanup_expired_locks
from io import BytesIO
from datetime import datetime

index_bp = Blueprint("index_bp", __name__)

@index_bp.route("/", endpoint="index")
@login_required
def index():
    # ページ件数指定: ?per_page=10/20/50/100/all
    per_page_raw = request.args.get('per_page', '20')
    if per_page_raw == 'all':
        per_page = None
        page = 1
        offset = 0
    else:
        per_page = int(per_page_raw)
        page = int(request.args.get('page', 1))
        offset = (page - 1) * per_page

    db = get_db()
    _cleanup_expired_locks(db)

    user_rows = db.execute(
        """
        SELECT username,
               COALESCE(NULLIF(realname, ''), username) AS display_name
        FROM users
        """
    ).fetchall()
    user_display = {r["username"]: r["display_name"] for r in user_rows}
    # realname(display_name) -> username への逆引き
    # ※ 同名が複数いる場合は最初の一致を使います（必要ならテンプレ側で username も表示して区別）
    display_to_username = {}
    for u, d in user_display.items():
        display_to_username.setdefault(d, u)

    logger.debug(user_display)

    # ===== フィルタ構築 =====
    filters = {}
    where = []
    params = []

    id_filter = request.args.get("id_filter", "").strip()
    filters["id"] = id_filter
    if id_filter:
        where.append("CAST(id AS TEXT) LIKE ?")
        params.append(f"%{id_filter}%")

    for f in INDEX_FIELDS:
        key = f["key"]
        v = request.args.get(f"{key}_filter", "").strip()

        # sample_manager は realname で受け取り、DB検索は username に正規化
        if key == "sample_manager" and v:
            normalized = display_to_username.get(v, v)  # realname -> username（見つからなければそのまま）
            filters[key] = v  # 画面反映は realname のまま
            where.append(f"{key} LIKE ?")
            params.append(f"%{normalized}%")
        else:
            filters[key] = v
            if v:
                where.append(f"{key} LIKE ?")
                params.append(f"%{v}%")

    sample_count_filter = request.args.get("sample_count_filter", "").strip()
    filters["sample_count"] = sample_count_filter

    where_clause = "WHERE " + " AND ".join(where) if where else ""

    # ===== データ取得（まず「全件」を取得して sample_count を計算）=====
    rows_all = db.execute(
        f"SELECT * FROM item {where_clause} ORDER BY id DESC",
        params
    ).fetchall()

    items_all = []
    for row in rows_all:
        item = dict(row)
        item_id = item["id"]
        child_total = db.execute(
            "SELECT COUNT(*) FROM child_item WHERE item_id=?",
            (item_id,)
        ).fetchone()[0]
        if child_total == 0:
            item["sample_count"] = item.get("num_of_samples", 0)
        else:
            cnt = db.execute(
                "SELECT COUNT(*) FROM child_item WHERE item_id=? AND status NOT IN (?, ?)",
                (item_id, "破棄", "譲渡")
            ).fetchone()[0]
            item["sample_count"] = cnt
        items_all.append(item)

    # sample_count フィルタ適用
    if sample_count_filter:
        items_filtered = [it for it in items_all if str(it["sample_count"]) == sample_count_filter]
    else:
        items_filtered = items_all

    # 合計件数 & ページング
    total = len(items_filtered)
    if per_page is not None:
        item_list = items_filtered[offset:offset + per_page]
    else:
        item_list = items_filtered

    # ===== フィルタ候補の辞書 =====
    filter_choices_dict = {}

    for f in INDEX_FIELDS:
        col = f["key"]
        rows = db.execute(
            f"SELECT DISTINCT {col} FROM item WHERE {col} IS NOT NULL AND {col} != ''"
        ).fetchall()

        # sample_manager だけは display_name（realname）で候補を返す
        if col == "sample_manager":
            usernames = {str(row[col]) for row in rows if row[col] not in (None, '')}
            # 表示名に変換（users に存在しない username はそのまま）
            display_names = {user_display.get(u, u) for u in usernames}
            filter_choices_dict[col] = sorted(display_names)
        else:
            filter_choices_dict[col] = sorted({str(row[col]) for row in rows if row[col] not in (None, '')})

    id_rows = db.execute("SELECT id FROM item ORDER BY id DESC").fetchall()
    filter_choices_dict["id"] = [str(r["id"]) for r in id_rows]

    # sample_count は「全件(items_all)」から候補を生成（ページング非依存）
    filter_choices_dict["sample_count"] = sorted({str(item["sample_count"]) for item in items_all}, key=lambda x: int(x) if x.isdigit() else x)

    # ===== ページ数 =====
    if per_page is not None:
        page_count = max(1, (total + per_page - 1) // per_page)
    else:
        page_count = 1

    return render_template(
        'index.html',
        items=item_list,
        page=page,
        page_count=page_count,
        filters=filters,
        total=total,
        fields=INDEX_FIELDS,
        filter_choices_dict=filter_choices_dict,
        per_page=per_page_raw,
        user_display=user_display,  # ラベル表示用（テンプレート側で sample_manager をこれで表示してもOK）
    )


@index_bp.route("/export_excel", methods=["GET"])
@login_required
def export_excel():
    """
    現在のフィルタ条件をそのまま適用した『全件』を Excel で出力する
    """
    db = get_db()
    _cleanup_expired_locks(db)

    # ===== users → 表示名マップ（index()と同じ）=====
    user_rows = db.execute(
        """
        SELECT username,
               COALESCE(NULLIF(realname, ''), username) AS display_name
        FROM users
        """
    ).fetchall()
    user_display = {r["username"]: r["display_name"] for r in user_rows}
    display_to_username = {}
    for u, d in user_display.items():
        display_to_username.setdefault(d, u)

    # ===== フィルタ構築（index()と同じロジック）=====
    filters = {}
    where = []
    params = []

    id_filter = request.args.get("id_filter", "").strip()
    filters["id"] = id_filter
    if id_filter:
        where.append("CAST(id AS TEXT) LIKE ?")
        params.append(f"%{id_filter}%")

    for f in INDEX_FIELDS:
        key = f["key"]
        v = request.args.get(f"{key}_filter", "").strip()

        if key == "sample_manager" and v:
            normalized = display_to_username.get(v, v)
            filters[key] = v
            where.append(f"{key} LIKE ?")
            params.append(f"%{normalized}%")
        else:
            filters[key] = v
            if v:
                where.append(f"{key} LIKE ?")
                params.append(f"%{v}%")

    sample_count_filter = request.args.get("sample_count_filter", "").strip()
    filters["sample_count"] = sample_count_filter

    where_clause = "WHERE " + " AND ".join(where) if where else ""

    # ===== データ取得（全件→sample_count計算→必要なら sample_count で絞り込み）=====
    rows_all = db.execute(
        f"SELECT * FROM item {where_clause} ORDER BY id DESC",
        params
    ).fetchall()

    items_all = []
    for row in rows_all:
        item = dict(row)
        item_id = item["id"]
        child_total = db.execute(
            "SELECT COUNT(*) FROM child_item WHERE item_id=?",
            (item_id,)
        ).fetchone()[0]
        if child_total == 0:
            item["sample_count"] = item.get("num_of_samples", 0)
        else:
            cnt = db.execute(
                "SELECT COUNT(*) FROM child_item WHERE item_id=? AND status NOT IN (?, ?)",
                (item_id, "破棄", "譲渡")
            ).fetchone()[0]
            item["sample_count"] = cnt
        items_all.append(item)

    if sample_count_filter:
        items_filtered = [it for it in items_all if str(it["sample_count"]) == sample_count_filter]
    else:
        items_filtered = items_all

    # ===== Excel 生成（openpyxl）=====
    # 依存: openpyxl（未導入なら `pip install openpyxl`）
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "items"

    # ヘッダ行: ID + INDEX_FIELDSの表示名 + サンプル数
    headers = ["ID"] + [f["name"] for f in INDEX_FIELDS] + ["サンプル数"]
    ws.append(headers)

    # データ行
    for it in items_filtered:
        row = [it["id"]]
        for f in INDEX_FIELDS:
            key = f["key"]
            val = it.get(key, "")
            if key == "sample_manager":
                val = user_display.get(val, val)
            row.append(val)
        row.append(it.get("sample_count", ""))
        ws.append(row)

    # ざっくり列幅調整（文字長の最大を元に）
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell_val = r[0].value
            if cell_val is None:
                continue
            max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # バイトストリームに保存して送信
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"items_{ts}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )